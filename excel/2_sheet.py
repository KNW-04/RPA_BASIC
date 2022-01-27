from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() # Create a new sheet with default name
ws.title = "Mysheet" # Change its name
ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("Sheet2222") # Create a new sheet and name it "Sheet2222"
''' Index Number '''
''' In This File, It's like
0 Sheet 1 Mysheet 2 Sheet2222 3 '''
ws2 = wb.create_sheet("NEW Sheet", 1) # locate it in the second place
''' Therefore it becomes like
0 Sheet 1 NEW Sheet 2 Mysheet 3 Sheet2222 4 '''

ws_new = wb["NEW Sheet"] # I didn't understand what it exactly does, but it seems to do the same function as line 12

print(wb.sheetnames) # Print all active sheets

# How to copy sheets
ws_new["A1"] = "test"
target = wb.copy_worksheet(ws_new) # copy
target.title = "copied sheet"

wb.save("sheet.xlsx")