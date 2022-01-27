from openpyxl import Workbook

wb =  Workbook()
ws = wb.active
ws.title = "TestSheet"

# input to a specified cell
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 5
ws["B2"] = 6
ws["B3"] = 7

# How to get Value of cells
print(ws["A1"]) # Print A1 cell's information
print(ws["A2"].value) # Returns A2 cell's value
print(ws["A100"].value) # Returns None if the cell is blank

# Or, You could try:
'''
ref:
row = 1, 2, 3, ...
column = A(1), B(2), C(3), ...

this format is useful when you need to operate with for ... or while ...
'''
print(ws.cell(row=1, column=1).value) # ws["A1"]'s Value

c = ws.cell(row=1, column=3, value=10) # ws["C1"].value = 10
print(c.value) # ws["C1"].value

ws_random = wb.create_sheet("Random", 0)

from random import randint
for i in range(1, 11):
    for k in range(1, 11):
        ws_random.cell(column=i, row=k, value=randint(1, 100000))


import schedule
mul_table = wb.create_sheet("multable")
idx = 0
def testfunc():
    print(f"{round(idx/1000000*100, 2)}% complete")
schedule.every(0.5).seconds.do(testfunc)

print("Writing values...")
for i in range(1, 1000):
    for k in range(1, 1000):
        idx += 1
        mul_table.cell(column=i, row=k, value=i*k)
        schedule.run_pending()
print("done")

print("Saving...")
wb.save("cell.xlsx")
print("Successfully saved the file")