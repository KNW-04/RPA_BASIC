from openpyxl import Workbook
from random import randint
wb = Workbook()
ws = wb.active

ws.append(["Idx", "sub", "Score"])

for i in range(1, 11):
    ws.append([i, "Math", randint(20, 100)])

# All rows
# print(tuple(ws.rows))
for row in tuple(ws.rows):
    print(row[2].value) # get the third row's value

# All columns
# print("\n\n")
# print(tuple(ws.columns))

for row in ws.iter_rows(): # All Rows
    print(row)  # You can use .value to get the cell's value

for col in ws.iter_cols(): # all columns
    print(col)

for row in ws.iter_rows(min_row=1, max_row=5): # Range
    print(row[2].value)

print()

# From row 1 to row 5, from col 2 to col 3
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row[0].value, row[1].value)
''' You can use the iter_cols to do the same thing, but in this case, it should be up/down wise '''

wb.save("6_cell_range_2.xlsx")