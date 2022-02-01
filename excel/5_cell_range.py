from openpyxl import Workbook
from random import randint
wb = Workbook()
ws = wb.active


ws.append(["No", "Eng", "Math"]) # import data from one by one line
for i in range(1, 11): # Total lines : 10
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # get "Eng" Column
for cell in col_B:
    print(cell.value)

col_range = ws["B:C"] # get datas from column B to C
for col in col_range:
    for cell in col:
        print(cell.value)

row_title = ws[1] # get the first row
for cell in row_title:
    print(cell.value)

row_range = ws[2:6] # get data from row 2 to 6 (Includes both endpoints)
for row in row_range:
    for cell in row:
        print(cell.value, end=" ")
    print()


from openpyxl.utils.cell import coordinate_from_string

for row in ws[1:ws.max_row]:
    for cell in row:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ") # A1 B1 C1 ...
        xy = coordinate_from_string(cell.coordinate)
        # print(xy, end=" ") # ('A', 1) ('B', 1) ('C', 1)
        print(xy[0], end="") # A
        print(xy[1], end=" ") # 1
    print()

wb.save("5_cell_range.xlsx")